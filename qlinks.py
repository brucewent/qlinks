#
# qlinks.py
# Test the links on a web site
# https://github.com/brucewent/qlinks is licensed under the MIT License
#

import sys, argparse, datetime
from urllib.parse import urljoin
import requests
from bs4 import BeautifulSoup
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

def main():
    # Parse the arguments, query the links, write out the results

    parser = argparse.ArgumentParser("qlinks - test the links on a web site")
    parser.add_argument("url", help="Input primary starting point URL, optional supporting URLS",
                        nargs="+")
    parser.add_argument("--recurse", dest='recurse', action='store_const',
                        const=True, default=False, help="Recurse through links on the site")
    parser.add_argument("--output", dest='output', type=str,
                        default="qlinks", help="Output file (default qlinks): \"<OUTPUT> YYYY-MM-DD.xlsx\"")
    args = parser.parse_args()

    # define the global variables

    global url_list, recursion_level, pages_done, urls_checked, data
    url_list = args.url
    recursion_level = -1
    pages_done = []
    urls_checked = []
    data = []

    start_time = datetime.datetime.now()
    print(start_time,"- qlinks starting")

    query_link(args.url[0], args.recurse)
    write_excel(args.output)

    finish_time = datetime.datetime.now()
    print(finish_time, "- qlinks finished")

    elapsed_time = finish_time - start_time
    print("Elapsed time", elapsed_time)

    return

def query_link (url, recurse):
    # retrieves the URL and examines the links on it

    global recursion_level, pages_done, urls_checked, data

    joe_blow = False

    recursion_level = recursion_level + 1
    prefix_space = recursion_level * "   " + str(recursion_level) + " "

    while True:
        # header string derived from a browser session
        headers = {"Pragma" : "no-cache",
            "Cache-Control" : "max-age=0",
            "Upgrade-Insecure-Requests" : "1",
            "User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36",
            "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "Referer" : get_base(url),
            "Accept-Encoding" : "gzip, deflate, br",
            "Accept-Language" : "en-US,en;q=0.9"}
        page_status = "OK"

        # read the referenced page
        try:
            r1 = requests.get(url, headers=headers, timeout=10)

        # catch all errors handled by requests
        except requests.exceptions.RequestException as e:
            page_status = "ConnectErr: " + str(e)
            pass

        # should catch the rest
        except Exception as e:
            page_status = "Unhandled: " + str(e)
            pass

        # return if the page was not available
        if page_status != "OK":
            try:
                if pages_done.index(url) >= 0:
                    recursion_level = recursion_level - 1
                    return
            except ValueError:
                pass
            pages_done.append(url)
            print(prefix_space + "Page", url, page_status)
            recursion_level = recursion_level - 1
            return

        # parse to a BeautifulSoup object
        soup = BeautifulSoup(r1.text, "lxml")
        try:
            page_title = soup.title.text
        except AttributeError:
            page_title = ""
        base_url = r1.url

        # quit if we did this before
        try:
            if pages_done.index(base_url) >= 0:
                recursion_level = recursion_level - 1
                joe_blow = True
                return
        except ValueError:
            pass

        # add it to the list of URLs done
        pages_done.append(base_url)

        print(prefix_space + "Page \"" + page_title + "\"(" + base_url + ")")

        # look for meta refresh redirects
        metas = soup.findAll('meta')
        redir_url = ""
        for meta in metas:
            try:
                if meta['http-equiv'] == "Refresh":
                    try:
                        content = meta['content']
                        p1 = content.find("=") + 1
                        if p1 > 0:
                            redir_url = content[p1:]
                        break
                    except KeyError:
                        pass
            except KeyError:
                pass
        url = urljoin(base_url, redir_url)

        # exit the forever loop
        if url == base_url:
            break
        print(prefix_space + "Redirecting...")

    # find all of the links and iterate over them
    links = soup.findAll('a')
    if not links:
        recursion_level = recursion_level - 1
        return

    links_checked = 0

    for a in links:
        try:
            link_href = a['href'].strip()
        except Exception as e:
            continue

        link_url = urljoin(base_url, link_href)

        if link_url.lower()[:4] != "http":
            continue

        link_status = 'OK'
        try:
            r2 = requests.head(link_url,headers=headers,timeout=10)

        # catches all errors handled by requests
        except requests.exceptions.RequestException as e:
            link_status = "ConnectErr: " + str(e)
            pass

        # should catch the rest
        except Exception as e:
            link_status = "Unhandled: " + str(e)
            pass

        # print links to increase console verbosity
        #print (prefix_space + "Link \"" + a.text.strip().replace('\n',' ') + "\" (" + link_url + ") " + link_status)

        # add it to the list

        data.append({ 'Page_Title' : page_title,
                      'Page_URL' : base_url,
                      'Link_Text' : a.text.strip().replace('\n',' '),
                      'Link_URL' : link_url,
                      'Link_Status' : link_status })

        links_checked = links_checked + 1

        # see if the link is in scope for recursion

        url_in_scope = False
        if get_base(base_url) == get_base(link_url):
            url_in_scope = True
        for i_url in url_list[1:]:
            if get_base(i_url) == get_base(link_url):
                url_in_scope = True

        # recurse if requested and the link is in scope

        if recurse and url_in_scope and link_status == 'OK' and \
                r2.headers['content-type'].find('text/html') > -1:
            query_link(link_url, recurse)

    recursion_level = recursion_level - 1
    return

def get_base(url):
    # Parses the base (domain) component of an URL

    my_url = url
    if my_url[len(my_url)-1:] != "/":
        my_url = my_url + "/"
    p1 = my_url.find("//")
    p2 = my_url.find("/", p1 + 2) + 1
    return my_url[:p2]

def write_excel(output):
    # writes the results to Excel

    global data
    if len(data) == 0:
        print ("No data to write to Excel")
        return

    # convert data (list of dictionaries) to df (pandas DataFrame)
    # and then Excel (openpyxl Workbook)

    df = DataFrame(data)[['Page_Title','Page_URL','Link_Text','Link_URL','Link_Status']]
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # pretty up the spreadsheet
    name_date = output + " " + str(datetime.datetime.now())[0:10]
    ws.title = name_date
    ws.freeze_panes = ws['A2']
    ws.auto_filter.ref = "A1:E" + str(df.shape[0]+1)

    # column width & label formatting
    cwidth = {'A' : 30, 'B' : 30, 'C' : 30, 'D' : 30, 'E' : 30}
    font = Font(b=True, i=True)
    fill = PatternFill(start_color='cccccc', end_color='cccccc', fill_type='solid')

    for xk in cwidth:
        ws.column_dimensions[xk].width = cwidth[xk]
        ws[xk+'1'].font = font
        ws[xk+'1'].fill = fill

    # set the hyperlinks
    for xr in range(df.shape[0]):
        ws['B'+str(xr+2)].hyperlink = ws['B'+str(xr+2)].value
        ws['B'+str(xr+2)].style = "Hyperlink"
        ws['D'+str(xr+2)].hyperlink = ws['D'+str(xr+2)].value
        ws['D'+str(xr+2)].style = "Hyperlink"

    # save & wrap up (loops to avoid file open errors)

    trys = 10
    fname = name_date
    for fnum in range (trys):
        try:
            wb.save(fname + ".xlsx")
            print ("Wrote", len(data), "lines to", fname + ".xlsx")
            break
        except Exception as e:
            if fnum+1 == trys:
                print(e)
            else:
                fname = name_date + "-" + str(fnum+1)
    return

# The following executes when run as a CLI

if __name__ == '__main__':
    main()
